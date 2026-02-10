import re
import json

from typing import List, Dict, Any, Optional

from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException

def extract_orig_req_from_excel_file(file_path: str, row_end:int) -> List[str]:
    results = []
    concatenated = []
    descriptions = []
    last_row_num = 0
    for row_num in range(2, row_end):
        main_requirement = get_requirement_from_excel(
            'ar_23/data.xlsx', 
            column_name='L1华为格式', 
            row_number=row_num
        )

        if main_requirement:
            results.append({
                'row': row_num,
                'req': main_requirement,
            })

    return results

def extract_content(text, field_name):
    """
    从遵循特定格式的字符串中提取指定字段的内容。

    预期的格式如下：
    【字段名1】
    字段1的内容...
    【字段名2】
    字段2的内容...

    Args:
        text (str): 包含格式化内容的字符串。
        field_name (str): 需要提取内容的字段名称。

    Returns:
        str: 指定字段的内容，如果未找到该字段则返回 None。
    """
    # 对字段名进行转义，以处理正则表达式中的特殊字符
    escaped_field_name = re.escape(field_name)
    
    # 正则表达式模式，用于查找指定字段与下一个字段之间或到字符串末尾的内容
    # re.DOTALL 使 '.' 可以匹配包括换行符在内的任意字符
    # (.*?) 是一个非贪婪匹配，用于捕获字段内容
    # 正向先行断言 (?=\n【|\Z) 确保匹配在下一个字段标记（前面有换行）或字符串末尾处停止
    pattern = f"【{escaped_field_name}】\s*(.*?)\s*(?=【|\Z)"
    
    match = re.search(pattern, text, re.DOTALL)
    
    if match:
        # 内容位于第一个捕获组中
        return match.group(1).strip()
    
    return None

def extract_descriptions_from_list(req_list: List[Dict[str, Any]]) -> List[str]:
    """
    从需求列表中提取所有description字段中的“需求描述”部分。
    """
    descriptions = []
    for req in req_list:
        if 'description' in req and req['description']:
            ar_description = extract_content(req['description'], "需求描述")
            if ar_description: # Make sure we got content
                descriptions.append(ar_description)
    return descriptions

def process_json_file(file_path: str) -> List[str]:
    """
    处理JSON文件，返回每个列表的description拼接结果
    """
    # 读取JSON文件
    with open(file_path, 'r', encoding='utf-8') as f:
        results_data = json.load(f)
    
    results = []
    
    # 遍历每个结果
    for item in results_data:
        row_n = item.get('row_number')
        decomposed_list = item.get('decomposed_list')

        if row_n and decomposed_list:
            descriptions = extract_descriptions_from_list(decomposed_list)
            
            # 拼接description
            concatenated = ' '.join(descriptions)
            
            results.append({
                'row': row_n,
                'concatenated': concatenated,
                'description_count': len(descriptions)
            })
    
    return results

def get_requirement_from_excel(file_path: str, column_name: str, row_number: int, sheet_name: Optional[str] = None) -> Optional[str]:
    """
    从Excel文件中通过列名和行号获取单元格内容。
    假设第一行为标题行。

    Args:
        file_path (str): Excel文件的路径。
        column_name (str): 目标列的名称（标题）。
        row_number (int): 目标行的编号（从1开始）。
        sheet_name (Optional[str], optional): 工作表的名称。默认为None，即活动工作表。

    Returns:
        Optional[str]: 读取到的需求文本或None。
    """
    try:
        workbook = load_workbook(filename=file_path)
        
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                print(f"[ERROR] 在文件 '{file_path}' 中未找到名为 '{sheet_name}' 的工作表。")
                return None
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active

        # 寻找列索引
        header = [cell.value for cell in sheet[1]]
        try:
            col_index = header.index(column_name) + 1
        except ValueError:
            print(f"[ERROR] 在工作表 '{sheet.title}' 的标题行中未找到列名 '{column_name}'。")
            return None

        # 检查行号是否有效
        if not (1 <= row_number <= sheet.max_row):
            print(f"[ERROR] 行号 {row_number} 超出有效范围 (1-{sheet.max_row})。")
            return None

        requirement = sheet.cell(row=row_number, column=col_index).value
        
        if requirement and isinstance(requirement, str):
            return requirement.strip()
        else:
            return None

    except FileNotFoundError:
        print(f"[ERROR] Excel文件未找到: {file_path}")
        return None
    except InvalidFileException:
        print(f"[ERROR] 无效的Excel文件格式或文件已损坏: {file_path}")
        return None
    except Exception as e:
        print(f"[ERROR] 读取Excel文件时发生未知错误: {e}")
        return None

def process_excel_file(file_path: str, row_end:int) -> List[str]:
    results = []
    concatenated = []
    descriptions = []
    last_row_num = 0
    for row_num in range(2, row_end):
        main_requirement = get_requirement_from_excel(
            'ar_23/data.xlsx', 
            column_name='L1华为格式', 
            row_number=row_num
        )

        if main_requirement:
            concatenated = ' '.join(descriptions)
            results.append({
                'row': last_row_num,
                'concatenated': concatenated,
                'description_count': len(descriptions)
            })
            concatenated = []
            descriptions = []
            last_row_num = row_num
        
        decomposed_requirement = get_requirement_from_excel(
            'ar_23/data.xlsx', 
            column_name='L2华为格式', 
            row_number=row_num
        )
        descriptions.append(extract_content(decomposed_requirement,"需求描述"))
        
    concatenated = ' '.join(descriptions)
    results.append({
        'row': last_row_num,
        'concatenated': concatenated,
        'description_count': len(descriptions)
    })
    return results
            

if __name__ == '__main__':
    res1 = process_json_file("ar_23\decomposed_output.json")
    with open("ar_23/ar_descriptions_1.json", 'w', encoding='utf-8') as f:
            json.dump(res1, f, ensure_ascii=False, indent=2)

    # res2 = process_excel_file("ar_23/data.xlsx",135)
    # with open("ar_23/ar_descriptions_ref.json", 'w', encoding='utf-8') as f:
    #     json.dump(res2, f, ensure_ascii=False, indent=2)

    # res3 = extract_orig_req_from_excel_file("ar_23/data.xlsx",135)
    # with open("ar_23/data.json", 'w', encoding='utf-8') as f:
    #     json.dump(res3, f, ensure_ascii=False, indent=2)


